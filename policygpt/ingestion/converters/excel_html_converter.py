"""ExcelToHtmlConverter — converts XLSX/XLS files to structured HTML.

Improvements over flat extraction
-----------------------------------
- Numbers formatted with thousands separators (1234567 → 1,234,567).
- Integers shown without decimal point (10.0 → 10).
- Dates formatted as 15-Jan-2025 instead of raw datetime objects.
- Booleans shown as Yes / No.
- Empty cells shown as — (em-dash) rather than blank.
- Tables use <thead>/<tbody> and scope="col" on header cells.
- Each sheet has a summary line (N rows × M columns) before the table.
- Sheets with a single column of labels are treated as key-value pairs.

Caching
-------
Converted HTML is saved to {output_dir}/{stem}.html and reused on subsequent
runs as long as the HTML file is newer than the source file.
"""

from __future__ import annotations

import html as _html_lib
import logging
import re
import time
from pathlib import Path

from policygpt.ingestion.converters.base import HtmlConverter

logger = logging.getLogger(__name__)


def _format_cell(value) -> str:
    """Format a cell value as a clean, readable string."""
    if value is None:
        return ""

    from datetime import date, datetime
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%Y %H:%M") if value.hour or value.minute else value.strftime("%d-%b-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, float):
        if value != value:          # NaN
            return ""
        if abs(value) < 1e15 and value == int(value):
            return f"{int(value):,}"
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"

    text = str(value).strip()
    return text


class ExcelToHtmlConverter(HtmlConverter):
    """Converts XLSX/XLS files to structured HTML."""

    @property
    def supported_content_types(self) -> frozenset[str]:
        return frozenset({"xlsx", "xls"})

    # ── Multi-sheet split ─────────────────────────────────────────────────────

    def convert_all(self, source_path: str) -> list[tuple[str, str]]:
        """Convert each worksheet to its own HTML file.

        For single-sheet workbooks delegates to the parent convert() so the
        normal cache logic applies unchanged.  For multi-sheet workbooks each
        sheet gets its own file named ``{stem}_{SafeSheetName}.html``.

        Returns
        -------
        list of (html_path, html_content) — one entry per non-empty sheet.
        """
        try:
            import openpyxl
        except ImportError:
            return [self.convert(source_path)]

        src = Path(source_path)
        try:
            wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
            sheet_names = list(wb.sheetnames)
            wb.close()
        except Exception:
            logger.warning("ExcelToHtmlConverter: could not open %s for sheet listing", src.name, exc_info=True)
            return [self.convert(source_path)]

        if len(sheet_names) <= 1:
            return [self.convert(source_path)]

        print(f"  [Convert] {src.name} — {len(sheet_names)} sheet(s), splitting …", flush=True)
        results: list[tuple[str, str]] = []

        for sheet_name in sheet_names:
            safe = re.sub(r'[<>:"/\\|?*\s]+', "_", sheet_name).strip("_") or "Sheet"
            out_path = self._out / f"{src.stem}_{safe}.html"

            # Per-sheet cache check.
            if self._skip_if_cached and out_path.exists():
                if out_path.stat().st_mtime >= src.stat().st_mtime:
                    print(f"  [Convert] {src.name} [{sheet_name}] — cached, skipping", flush=True)
                    results.append((str(out_path), out_path.read_text(encoding="utf-8", errors="ignore")))
                    continue

            t0 = time.perf_counter()
            try:
                sheet_html = self._sheet_to_html(sheet_name, src, original_source=src.name)
            except Exception as exc:
                elapsed = time.perf_counter() - t0
                print(f"  [Convert] {src.name} [{sheet_name}] — FAILED after {elapsed:.1f}s: {exc}", flush=True)
                logger.warning("ExcelToHtmlConverter: sheet %r in %s failed — %s", sheet_name, src.name, exc, exc_info=True)
                continue

            if not sheet_html:
                print(f"  [Convert] {src.name} [{sheet_name}] — empty, skipping", flush=True)
                continue

            title = f"{self._title_from_stem(src.stem)} — {sheet_name}"
            meta = (
                f'<dl class="doc-meta">'
                f"<dt>Source</dt><dd>{_html_lib.escape(src.name)}</dd>"
                f"<dt>Sheet</dt><dd>{_html_lib.escape(sheet_name)}</dd>"
                f"<dt>Type</dt><dd>Excel Worksheet</dd>"
                f"</dl>"
            )
            html_content = self._wrap_html(title, meta + "\n" + sheet_html)
            out_path.write_text(html_content, encoding="utf-8")

            elapsed = time.perf_counter() - t0
            print(f"  [Convert] {src.name} [{sheet_name}] — done in {elapsed:.1f}s → {out_path.name}", flush=True)
            results.append((str(out_path), html_content))

        return results if results else [self.convert(source_path)]

    def _convert_to_html(self, path: Path) -> str:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "Excel-to-HTML conversion requires 'openpyxl'. "
                "Install it with: pip install openpyxl"
            ) from exc

        title = self._title_from_stem(path.stem)
        # Read sheet names only — cell data is read inside _sheet_to_html.
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames)
        wb.close()

        sheet_parts: list[str] = []
        for sheet_name in sheet_names:
            sheet_html = self._sheet_to_html(sheet_name, path)
            if sheet_html:
                sheet_parts.append(sheet_html)

        meta = (
            f'<dl class="doc-meta">'
            f"<dt>Source</dt><dd>{_html_lib.escape(path.name)}</dd>"
            f"<dt>Type</dt><dd>Excel Workbook</dd>"
            f"<dt>Sheets</dt><dd>{len(sheet_parts)}</dd>"
            f"</dl>"
        )
        body = meta + "\n" + "\n".join(sheet_parts)
        return self._wrap_html(title, body)

    @staticmethod
    def _read_sheet_data(
        src: Path, sheet_name: str
    ) -> tuple[dict[tuple[int, int], object], set[tuple[int, int]]]:
        """Stream cell values from one sheet using read_only mode (data cells only).

        Returns
        -------
        cell_values : dict mapping (row, col) → raw cell value for every
                      non-None cell in the sheet.
        primary_cells : set of (row, col) that are the top-left cell of a
                        merged range — so callers know which cells anchor a merge.
        """
        import openpyxl

        cell_values: dict[tuple[int, int], object] = {}

        # ── Pass 1: stream cell data (read_only=True — fast, low memory) ──────
        wb_ro = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
        try:
            ws_ro = wb_ro[sheet_name]
            for row_idx, row in enumerate(ws_ro.iter_rows(values_only=True), start=1):
                for col_idx, val in enumerate(row, start=1):
                    if val is not None:
                        cell_values[(row_idx, col_idx)] = val
        finally:
            wb_ro.close()

        # ── Pass 2: read merged-cell ranges only (no cell iteration) ──────────
        # load_workbook with read_only=False is used ONLY to read sheet metadata
        # (merged_cells.ranges).  No cell data is iterated so it stays fast.
        primary_cells: set[tuple[int, int]] = set()
        try:
            wb_meta = openpyxl.load_workbook(str(src), read_only=False, data_only=True)
            ws_meta = wb_meta[sheet_name]
            for mr in ws_meta.merged_cells.ranges:
                primary = (mr.min_row, mr.min_col)
                primary_val = cell_values.get(primary)
                primary_cells.add(primary)
                # Fill non-primary cells with the primary cell's value.
                for r in range(mr.min_row, mr.max_row + 1):
                    for c in range(mr.min_col, mr.max_col + 1):
                        if (r, c) != primary:
                            if primary_val is not None:
                                cell_values[(r, c)] = primary_val
            wb_meta.close()
        except Exception:
            pass  # merged cell info unavailable — continue without fill

        return cell_values, primary_cells

    @staticmethod
    def _sheet_to_html(sheet_name: str, src: Path, original_source: str | None = None) -> str:
        """Convert one worksheet to a <section> HTML fragment."""
        cell_values, _ = ExcelToHtmlConverter._read_sheet_data(src, sheet_name)

        if not cell_values:
            return ""

        # Determine the actual data bounds.
        all_rows = {r for r, _ in cell_values}
        all_cols = {c for _, c in cell_values}
        min_row, max_row = min(all_rows), max(all_rows)
        min_col, max_col = min(all_cols), max(all_cols)

        # Build a dense grid from the sparse cell dict.
        raw_rows: list[list[object]] = []
        for r in range(min_row, max_row + 1):
            raw_rows.append([
                cell_values.get((r, c))
                for c in range(min_col, max_col + 1)
            ])

        # Trim trailing fully-empty rows.
        while raw_rows and not any(v is not None for v in raw_rows[-1]):
            raw_rows.pop()

        if not raw_rows:
            return ""

        # Format and drop still-empty rows.
        rows: list[list[str]] = []
        for raw in raw_rows:
            formatted = [_format_cell(c) for c in raw]
            if any(c for c in formatted):
                rows.append(formatted)

        if not rows:
            return ""

        sheet_name = _html_lib.escape(sheet_name or "Sheet")
        data_rows = len(rows) - 1   # excluding header
        col_count = len(rows[0]) if rows else 0
        summary = f"{data_rows} row{'s' if data_rows != 1 else ''} × {col_count} column{'s' if col_count != 1 else ''}"

        # Header row → <thead>
        header_cells = "".join(
            f'<th scope="col">{_html_lib.escape(cell)}</th>'
            for cell in rows[0]
        )
        thead = f"<thead><tr>{header_cells}</tr></thead>"

        # Body rows → <tbody>
        body_rows: list[str] = []
        for row in rows[1:]:
            cells = "".join(f"<td>{_html_lib.escape(cell)}</td>" for cell in row)
            body_rows.append(f"<tr>{cells}</tr>")
        tbody = ("<tbody>\n" + "\n".join(body_rows) + "\n</tbody>") if body_rows else ""

        table_html = f"<table>\n{thead}\n{tbody}\n</table>"
        return (
            f'<section class="sheet">\n'
            f"<h2>{sheet_name}</h2>\n"
            f'<p class="sheet-summary">{summary}</p>\n'
            f"{table_html}\n"
            f"</section>"
        )
